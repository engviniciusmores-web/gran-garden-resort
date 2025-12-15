import openpyxl
from datetime import datetime

# Ler arquivo
wb = openpyxl.load_workbook('cronograma_gav-gran-garden_1765606540.xlsx', data_only=True)
ws = wb['Activities']

# Colunas importantes (baseado no cabeçalho linha 4):
# 1: ID
# 2: Pacote de trabalho/tarefas
# 4: Grupo de replicação (BLOCO)
# 5: Lote (nível)
# 8: Data início linha base
# 9: Data término linha base
# 10: Data Início real
# 11: Data Término real
# 22: Base (%)
# 23: Previsto (%)
# 24: Realizado (%)
# 38: Término real

tasks_data = []
for row_idx in range(5, min(ws.max_row + 1, 200)):  # Processar primeiras 200 linhas
    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    
    task_name = str(row[2]) if row[2] else ''
    bloco = str(row[4]) if row[4] else ''
    lote = str(row[5]) if row[5] else ''
    start_baseline = row[8]
    end_baseline = row[9]
    start_real = row[10]
    end_real = row[11]
    base_pct = row[22]
    previsto_pct = row[23]
    realizado_pct = row[24]
    
    # Filtrar apenas tarefas relevantes dos blocos principais
    if any(keyword in task_name.lower() for keyword in ['fundação', 'estrutura', 'alvenaria', 'limpeza', 'escavações', 'contenções', 'instalações', 'revestimento']):
        if bloco and bloco not in ['Não agrupado', '-'] and lote:
            
            # Formatar datas
            def format_date(dt):
                if isinstance(dt, datetime):
                    return dt.strftime('%Y-%m-%d')
                return ''
            
            # Determinar status
            status = 'A Fazer'
            try:
                realizado_val = float(realizado_pct) if realizado_pct else 0
                if realizado_val == 100:
                    status = 'Concluído'
                elif realizado_val > 0:
                    status = 'Em Andamento'
            except:
                realizado_val = 0
            
            tasks_data.append({
                'name': f'{task_name} {bloco} - {lote}',
                'bloco': bloco,
                'lote': lote,
                'start_baseline': format_date(start_baseline),
                'end_baseline': format_date(end_baseline),
                'start_real': format_date(start_real),
                'end_real': format_date(end_real),
                'base_pct': base_pct if base_pct else 0,
                'previsto_pct': previsto_pct if previsto_pct else 0,
                'realizado_pct': realizado_val,
                'status': status
            })

# Mostrar top 30 tarefas por prioridade (já executadas primeiro)
tasks_data.sort(key=lambda x: (x['status'] != 'Concluído', x['status'] != 'Em Andamento', x['start_baseline']))

print("="*120)
print("TOP 30 TAREFAS PARA O CONSTANTS.TS")
print("="*120)

for i, task in enumerate(tasks_data[:30], 1):
    print(f"\n{i}. {task['name']}")
    print(f"   Status: {task['status']} | Realizado: {task['realizado_pct']}%")
    print(f"   Planejado: {task['start_baseline']} → {task['end_baseline']}")
    if task['start_real']:
        print(f"   Real: {task['start_real']} → {task['end_real'] if task['end_real'] else 'Em andamento'}")

# Gerar código TypeScript
print("\n\n" + "="*120)
print("CÓDIGO TYPESCRIPT PARA CONSTANTS.TS")
print("="*120)

print("\nexport const MOCK_TASKS: Task[] = [")

for i, task in enumerate(tasks_data[:20], 1):
    # Determinar equipe baseada no bloco
    if 'A1' in task['bloco'] or 'A2' in task['bloco'] or 'A3' in task['bloco'] or 'A4' in task['bloco']:
        responsible = 'Equipe Caio'
    elif 'B' in task['bloco']:
        responsible = 'Equipe Lucas'  
    elif 'C' in task['bloco']:
        responsible = 'Equipe Lucas'
    elif 'Lazer' in task['bloco'] or 'Recepção' in task['bloco']:
        responsible = 'Equipe Tatiane'
    else:
        responsible = 'Coord. Geral'
    
    print(f"  {{ ")
    print(f"    id: '{i}', ")
    print(f"    name: '{task['name']}', ")
    print(f"    deadline: '{task['end_baseline']}', ")
    print(f"    status: '{task['status']}', ")
    print(f"    responsible: '{responsible}',")
    print(f"    plannedStart: '{task['start_baseline']}',")
    print(f"    plannedEnd: '{task['end_baseline']}',")
    print(f"    actualStart: '{task['start_real']}' || undefined,")
    print(f"    actualEnd: '{task['end_real']}' || undefined,")
    print(f"    progress: {int(task['realizado_pct'])}")
    print(f"  }},")

print("];")
