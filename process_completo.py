import openpyxl
from datetime import datetime
import json

print("="*120)
print("PROCESSANDO TODO O CRONOGRAMA GAV - GRAN GARDEN RESORT")
print("="*120)

# Ler arquivo Excel
wb = openpyxl.load_workbook('cronograma_gav-gran-garden_1765606540.xlsx', data_only=True)
ws = wb['Activities']

print(f"\nTotal de linhas no Excel: {ws.max_row}")

# Processar TODAS as linhas
all_tasks = []
tasks_by_bloco = {}
tasks_by_service = {}

for row_idx in range(5, ws.max_row + 1):
    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    
    task_id = row[1]
    task_name = str(row[2]) if row[2] else ''
    servico = str(row[3]) if row[3] else '-'
    bloco = str(row[4]) if row[4] else ''
    lote = str(row[5]) if row[5] else ''
    start_baseline = row[8]
    end_baseline = row[9]
    start_real = row[10]
    end_real = row[11]
    duration = row[12]
    responsavel = str(row[14]) if row[14] else '-'
    base_pct = row[22]
    previsto_pct = row[23]
    realizado_pct = row[24]
    inicio_real = row[38]
    termino_real = row[39]
    
    # Filtrar tarefas relevantes (não vazias)
    if task_name and task_name not in ['-', 'None'] and bloco and bloco not in ['Não agrupado', '-']:
        
        def format_date(dt):
            if isinstance(dt, datetime):
                return dt.strftime('%Y-%m-%d')
            return ''
        
        # Converter percentuais
        try:
            realizado_val = float(realizado_pct) if realizado_pct else 0
            previsto_val = float(previsto_pct) if previsto_pct else 0
            base_val = float(base_pct) if base_pct else 0
        except:
            realizado_val = 0
            previsto_val = 0
            base_val = 0
        
        # Determinar status detalhado
        if realizado_val == 100:
            status = 'Concluído'
        elif realizado_val >= 75:
            status = 'Em Andamento - Avançado'
        elif realizado_val >= 50:
            status = 'Em Andamento'
        elif realizado_val > 0:
            status = 'Em Andamento - Inicial'
        elif previsto_val > 0:
            status = 'Atrasado'
        else:
            status = 'A Fazer'
        
        task_data = {
            'id': task_id,
            'name': task_name,
            'servico': servico,
            'bloco': bloco,
            'lote': lote,
            'start_baseline': format_date(start_baseline),
            'end_baseline': format_date(end_baseline),
            'start_real': format_date(start_real),
            'end_real': format_date(end_real),
            'duration': duration if duration else 0,
            'responsavel': responsavel,
            'base_pct': base_val,
            'previsto_pct': previsto_val,
            'realizado_pct': realizado_val,
            'status': status,
            'inicio_real': format_date(inicio_real) if inicio_real else '',
            'termino_real': format_date(termino_real) if termino_real else '',
        }
        
        all_tasks.append(task_data)
        
        # Agrupar por bloco
        if bloco not in tasks_by_bloco:
            tasks_by_bloco[bloco] = []
        tasks_by_bloco[bloco].append(task_data)
        
        # Agrupar por serviço
        if task_name not in tasks_by_service:
            tasks_by_service[task_name] = []
        tasks_by_service[task_name].append(task_data)

print(f"\nTotal de tarefas extraídas: {len(all_tasks)}")
print(f"Blocos encontrados: {len(tasks_by_bloco)}")
print(f"Tipos de serviço: {len(tasks_by_service)}")

# Estatísticas por bloco
print("\n" + "="*120)
print("ESTATÍSTICAS POR BLOCO")
print("="*120)

for bloco in sorted(tasks_by_bloco.keys()):
    tasks = tasks_by_bloco[bloco]
    concluidas = sum(1 for t in tasks if t['status'] == 'Concluído')
    em_andamento = sum(1 for t in tasks if 'Em Andamento' in t['status'])
    atrasadas = sum(1 for t in tasks if t['status'] == 'Atrasado')
    a_fazer = sum(1 for t in tasks if t['status'] == 'A Fazer')
    avg_progress = sum(t['realizado_pct'] for t in tasks) / len(tasks) if tasks else 0
    
    print(f"\n{bloco}:")
    print(f"  Total: {len(tasks)} tarefas")
    print(f"  Concluídas: {concluidas} ({concluidas/len(tasks)*100:.1f}%)")
    print(f"  Em Andamento: {em_andamento} ({em_andamento/len(tasks)*100:.1f}%)")
    print(f"  Atrasadas: {atrasadas} ({atrasadas/len(tasks)*100:.1f}%)")
    print(f"  A Fazer: {a_fazer} ({a_fazer/len(tasks)*100:.1f}%)")
    print(f"  Progresso Médio: {avg_progress:.1f}%")

# Principais serviços
print("\n" + "="*120)
print("TIPOS DE SERVIÇO MAIS COMUNS")
print("="*120)

service_counts = {k: len(v) for k, v in tasks_by_service.items()}
top_services = sorted(service_counts.items(), key=lambda x: x[1], reverse=True)[:20]

for service, count in top_services:
    tasks = tasks_by_service[service]
    avg_progress = sum(t['realizado_pct'] for t in tasks) / len(tasks)
    print(f"{service}: {count} tarefas - {avg_progress:.1f}% concluído")

# Salvar todas as tarefas em JSON
with open('todas_tarefas.json', 'w', encoding='utf-8') as f:
    json.dump(all_tasks, f, ensure_ascii=False, indent=2)

print(f"\n✅ Dados salvos em: todas_tarefas.json")

# Gerar código TypeScript com TODAS as tarefas importantes
print("\n" + "="*120)
print("GERANDO CÓDIGO TYPESCRIPT COMPLETO")
print("="*120)

# Selecionar tarefas mais importantes: concluídas, em andamento e críticas
important_tasks = []

# 1. Todas concluídas (para histórico)
concluidas = [t for t in all_tasks if t['status'] == 'Concluído']
important_tasks.extend(concluidas[:30])  # Top 30 concluídas

# 2. Todas em andamento
em_andamento = [t for t in all_tasks if 'Em Andamento' in t['status']]
important_tasks.extend(em_andamento)

# 3. Atrasadas
atrasadas = [t for t in all_tasks if t['status'] == 'Atrasado']
important_tasks.extend(atrasadas[:15])

# 4. Próximas (A Fazer com data próxima)
a_fazer = [t for t in all_tasks if t['status'] == 'A Fazer' and t['start_baseline']]
a_fazer_sorted = sorted(a_fazer, key=lambda x: x['start_baseline'])[:25]
important_tasks.extend(a_fazer_sorted)

# Remover duplicatas mantendo ordem
seen = set()
unique_tasks = []
for task in important_tasks:
    if task['id'] not in seen:
        seen.add(task['id'])
        unique_tasks.append(task)

print(f"\nTotal de tarefas selecionadas para o sistema: {len(unique_tasks)}")

# Gerar TypeScript
ts_output = []
ts_output.append("// DADOS COMPLETOS DO CRONOGRAMA GAV - Gran Garden Resort")
ts_output.append("// Gerado automaticamente em 13/12/2025")
ts_output.append("// Total de tarefas no cronograma: " + str(len(all_tasks)))
ts_output.append("")
ts_output.append("export const MOCK_TASKS: Task[] = [")

for i, task in enumerate(unique_tasks[:100], 1):  # Limitar a 100 para não sobrecarregar
    # Determinar equipe
    if any(x in task['bloco'] for x in ['A1', 'A2', 'A3', 'A4']):
        responsible = 'Equipe Caio'
    elif any(x in task['bloco'] for x in ['B1', 'B2', 'B3', 'B4']):
        responsible = 'Equipe Lucas'
    elif any(x in task['bloco'] for x in ['C1', 'C2', 'C3', 'C4']):
        responsible = 'Equipe Lucas'
    elif 'Lazer' in task['bloco'] or 'Recep' in task['bloco']:
        responsible = 'Equipe Tatiane'
    else:
        responsible = task['responsavel'] if task['responsavel'] != '-' else 'Coord. Geral'
    
    # Montar nome completo
    full_name = f"{task['name']} {task['bloco']} - {task['lote']}"
    
    ts_output.append("  {")
    ts_output.append(f"    id: '{i}',")
    ts_output.append(f"    name: '{full_name}',")
    ts_output.append(f"    deadline: '{task['end_baseline']}',")
    ts_output.append(f"    status: '{task['status']}',")
    ts_output.append(f"    responsible: '{responsible}',")
    ts_output.append(f"    plannedStart: '{task['start_baseline']}',")
    ts_output.append(f"    plannedEnd: '{task['end_baseline']}',")
    ts_output.append(f"    actualStart: '{task['start_real']}' || undefined,")
    ts_output.append(f"    actualEnd: '{task['end_real']}' || undefined,")
    ts_output.append(f"    progress: {int(task['realizado_pct'])}")
    ts_output.append("  },")

ts_output.append("];")

# Salvar TypeScript
with open('constants_completo.ts', 'w', encoding='utf-8') as f:
    f.write('\n'.join(ts_output))

print(f"\n✅ Código TypeScript salvo em: constants_completo.ts")

# GERAR RELATÓRIOS EM PDF E WORD
print("\n" + "="*120)
print("GERANDO RELATÓRIOS PROFISSIONAIS")
print("="*120)

from report_generator import generate_report_from_json

# Preparar dados do relatório
report_data = {
    'type': 'geral',
    'project': 'Gran Garden Resort - Cronograma Completo',
    'period': 'all-time',
    'period_label': 'Todo o Período',
    'tasks': len(all_tasks),
    'completedTasks': len([t for t in all_tasks if t['status'] == 'Concluído']),
    'includeCharts': True,
    'includePhotos': False,
    'generatedAt': datetime.now().isoformat()
}

# Gerar PDF
try:
    pdf_path = generate_report_from_json(report_data, format='pdf')
    print(f"\n✅ Relatório PDF gerado: {pdf_path}")
except Exception as e:
    print(f"\n⚠️  Erro ao gerar PDF: {e}")

# Gerar Word
try:
    word_path = generate_report_from_json(report_data, format='word')
    print(f"✅ Relatório Word gerado: {word_path}")
except Exception as e:
    print(f"⚠️  Erro ao gerar Word: {e}")

print(f"\n📊 Resumo Final:")
print(f"   - Total extraído: {len(all_tasks)} tarefas")
print(f"   - Selecionadas: {len(unique_tasks)} tarefas importantes")
print(f"   - Exportadas: {min(100, len(unique_tasks))} para o sistema")
print(f"   - Relatórios gerados em: ./relatorios/")
