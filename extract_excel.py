import openpyxl
from datetime import datetime
import re

# Ler o arquivo Excel
file_path = 'cronograma_gav-gran-garden_1765606540.xlsx'

print("Carregando arquivo Excel...")
wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"Abas encontradas: {wb.sheetnames}\n")

# Processar cada aba
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*100}")
    print(f"ABA: {sheet_name}")
    print('='*100)
    
    # Mostrar primeiras 50 linhas
    for row_idx, row in enumerate(ws.iter_rows(max_row=50, values_only=True), 1):
        # Filtrar linhas que contenham informações úteis (datas, atividades)
        row_str = str(row)
        if any(keyword in row_str.lower() for keyword in ['fundação', 'estrutura', 'alvenaria', 'instalações', 'revestimento', 'pavimento', 'térreo', 'laje', 'limpeza', 'escavação', 'contenção']):
            print(f"L{row_idx}: {row}")
        elif row_idx <= 10:  # Sempre mostrar primeiras 10 linhas (cabeçalhos)
            print(f"L{row_idx}: {row}")
            
print("\n\n" + "="*100)
print("RESUMO: Buscando colunas de datas e atividades...")
print("="*100)

# Tentar identificar estrutura do cronograma na primeira aba útil
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n\nAnalisando estrutura da aba: {sheet_name}")
    
    # Procurar linha de cabeçalho
    for row_idx in range(1, 20):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        row_str = ' '.join([str(cell).lower() for cell in row if cell])
        
        if 'atividade' in row_str or 'tarefa' in row_str or 'descrição' in row_str:
            print(f"\nPossível cabeçalho na linha {row_idx}:")
            print(row)
            
            # Mostrar próximas 30 linhas de dados
            print(f"\nDados a partir da linha {row_idx + 1}:")
            for data_row_idx in range(row_idx + 1, min(row_idx + 31, ws.max_row + 1)):
                data_row = list(ws.iter_rows(min_row=data_row_idx, max_row=data_row_idx, values_only=True))[0]
                print(f"L{data_row_idx}: {data_row}")
            break
